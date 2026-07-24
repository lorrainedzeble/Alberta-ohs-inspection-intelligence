"""Download and validate the current Alberta Employer Records workbook.

The catalogue lookup keeps the public repository free of a 53 MB binary while
preserving a stable local filename for the analytical skills.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
import urllib.request
import zipfile
from pathlib import Path

CATALOGUE_API = (
    "https://open.canada.ca/data/api/action/package_show"
    "?id=a2772d8c-48be-4d39-bcf2-dafca456d724"
)
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "2024_ohs-employer-record-open-data.xlsx"
REQUIRED_SHEETS = {
    "Injury (2020-2024)",
    "Order (2020-2024)",
    "Penalty (2020-2024)",
    "Ticket (2020-2024)",
    "Investigation (2020-2024)",
    "Acceptance (2020-2024)",
    "Approval (2020-2024)",
    "Conviction (2020-2024)",
}


def _request_json(url: str) -> dict:
    request = urllib.request.Request(url, headers={"User-Agent": "ohs-inspection-intelligence/1.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.load(response)


def resolve_xlsx_url() -> tuple[str, dict]:
    payload = _request_json(CATALOGUE_API)
    if not payload.get("success"):
        raise RuntimeError("Open-data catalogue returned success=false.")
    resources = payload["result"].get("resources", [])
    candidates = [
        resource
        for resource in resources
        if str(resource.get("format", "")).upper() == "XLSX"
        and "employer" in str(resource.get("name", "")).lower()
    ]
    if not candidates:
        raise RuntimeError("No Employer Records XLSX resource was found in the catalogue.")
    resource = max(candidates, key=lambda item: item.get("last_modified") or item.get("created") or "")
    return resource["url"], resource


def download(url: str, target: Path) -> tuple[str, int]:
    target.parent.mkdir(parents=True, exist_ok=True)
    partial = target.with_suffix(target.suffix + ".part")
    digest = hashlib.sha256()
    total = 0
    request = urllib.request.Request(url, headers={"User-Agent": "ohs-inspection-intelligence/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=120) as response, partial.open("wb") as output:
            while chunk := response.read(1024 * 1024):
                output.write(chunk)
                digest.update(chunk)
                total += len(chunk)
                print(f"\rDownloaded {total / 1048576:.1f} MB", end="", file=sys.stderr, flush=True)
        print(file=sys.stderr)
        os.replace(partial, target)
    finally:
        if partial.exists():
            partial.unlink()
    return digest.hexdigest(), total


def validate_xlsx(path: Path) -> list[str]:
    if not zipfile.is_zipfile(path):
        raise RuntimeError("Downloaded file is not a valid XLSX/ZIP container.")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    finally:
        workbook.close()
    if missing:
        raise RuntimeError(f"Workbook schema drift: missing sheets {missing}")
    return sorted(REQUIRED_SHEETS)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", help="Override the catalogue-resolved XLSX URL.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--force", action="store_true", help="Replace an existing local workbook.")
    args = parser.parse_args()

    output = args.output.resolve()
    if output.exists() and not args.force:
        print(f"{output} already exists. Use --force to refresh it.", file=sys.stderr)
        return 2

    url, resource = (args.url, {"name": "manual override"}) if args.url else resolve_xlsx_url()
    started = time.time()
    sha256, size = download(url, output)
    sheets = validate_xlsx(output)
    receipt = {
        "downloaded_at_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "source_url": url,
        "catalogue_resource": resource,
        "local_path": str(output.relative_to(PROJECT_ROOT)) if output.is_relative_to(PROJECT_ROOT) else str(output),
        "bytes": size,
        "sha256": sha256,
        "validated_sheets": sheets,
        "elapsed_seconds": round(time.time() - started, 1),
    }
    receipt_path = output.parent / "refresh-receipt.json"
    receipt_path.write_text(json.dumps(receipt, indent=2), encoding="utf-8")
    print(json.dumps(receipt, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
